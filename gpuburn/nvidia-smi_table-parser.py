# nvidia-smi_table-parser.py
# requirements.txt >> pandas>=2.0.0 openpyxl>=3.1.0
#!/usr/bin/env python3
import argparse
import re
from typing import Dict, Any, List, Optional
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

TIMESTAMP_RE = re.compile(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})$")
GPU_INFO_RE = re.compile(r"^\|\s+(\d+)\s+\S")
SEPARATOR_RE = re.compile(r"^\+\-+|\|\=+|\|\s*$")

def parse_metrics_line(line: str) -> Dict[str, Optional[int]]:
    parts = [p.strip() for p in line.split("|")]
    metrics: Dict[str, Optional[int]] = {
        "temp_c": None,
        "power_w": None,
        "power_capacity_w": None,
        "memory_used_mib": None,
        "memory_total_mib": None,
        "utilization": None,
    }
    if len(parts) < 4:
        return metrics

    col0 = parts[1]
    m_temp = re.search(r"(\d+)C", col0)
    if m_temp:
        metrics["temp_c"] = int(m_temp.group(1))
    m_pwr = re.search(r"(\d+)W\s*/\s*(\d+)W", col0)
    if m_pwr:
        metrics["power_w"] = int(m_pwr.group(1))
        metrics["power_capacity_w"] = int(m_pwr.group(2))

    col1 = parts[2]
    m_mem = re.search(r"(\d+)MiB\s*/\s*(\d+)MiB", col1)
    if m_mem:
        metrics["memory_used_mib"] = int(m_mem.group(1))
        metrics["memory_total_mib"] = int(m_mem.group(2))

    col2 = parts[3]
    m_util = re.search(r"(\d+)%", col2)
    if m_util:
        metrics["utilization"] = int(m_util.group(1))

    return metrics

def parse_log(filepath: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    current_ts: Optional[str] = None
    current_snapshot: Dict[str, Any] = {}
    expecting_metrics_for_gpu: Optional[int] = None

    def flush_snapshot():
        nonlocal current_snapshot
        if not current_snapshot:
            return
        row: Dict[str, Any] = {"timestamp": current_snapshot.get("timestamp")}
        # Переносим все собранные по снапшоту метрики любых GPU
        for key, metrics in current_snapshot.items():
            if not key.startswith("gpu"):
                continue
            row[f"{key}_temp_c"] = metrics.get("temp_c")
            row[f"{key}_power_w"] = metrics.get("power_w")
            row[f"{key}_power_capacity_w"] = metrics.get("power_capacity_w")
            row[f"{key}_memory_used_mib"] = metrics.get("memory_used_mib")
            row[f"{key}_memory_total_mib"] = metrics.get("memory_total_mib")
            row[f"{key}_utilization"] = metrics.get("utilization")
        rows.append(row)
        current_snapshot = {}

    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        for raw in f:
            line = raw.rstrip("\n")

            m_ts = TIMESTAMP_RE.match(line)
            if m_ts:
                if current_ts is not None:
                    flush_snapshot()
                current_ts = m_ts.group(1)
                current_snapshot = {"timestamp": current_ts}
                expecting_metrics_for_gpu = None
                continue

            m_gpu = GPU_INFO_RE.match(line)
            if m_gpu:
                try:
                    expecting_metrics_for_gpu = int(m_gpu.group(1))
                except ValueError:
                    expecting_metrics_for_gpu = None
                continue

            if SEPARATOR_RE.match(line):
                continue

            if expecting_metrics_for_gpu is not None:
                current_snapshot[f"gpu{expecting_metrics_for_gpu}"] = parse_metrics_line(line)
                expecting_metrics_for_gpu = None
                continue

    if current_ts is not None and current_snapshot:
        flush_snapshot()

    return rows

def main():
    ap = argparse.ArgumentParser(description="Parse nvidia-smi log to Excel for any number of GPUs.")
    ap.add_argument("input", help="Path to nvidia_smi.log")
    ap.add_argument("-o", "--output", default="gpu_metrics.xlsx", help="Output .xlsx file")
    ap.add_argument("--sheet", default="metrics", help="Sheet name")
    args = ap.parse_args()

    rows = parse_log(args.input)
    if not rows:
        print("No data parsed.")
        return

    df = pd.DataFrame(rows)

    # Динамическое формирование колонок под N GPU
    # Находим все встреченные индексы GPU из имен колонок вида gpu{N}_metric
    gpu_ids: List[int] = []
    for col in df.columns:
        if col.startswith("gpu") and "_" in col:
            try:
                idx = int(col[3:col.index("_")])
                if idx not in gpu_ids:
                    gpu_ids.append(idx)
            except Exception:
                pass
    gpu_ids.sort()

    metric_suffixes = [
        "temp_c",
        "power_w",
        "power_capacity_w",
        "memory_used_mib",
        "memory_total_mib",
        "utilization",
    ]

    columns = ["timestamp"]
    for gid in gpu_ids:
        for suffix in metric_suffixes:
            cname = f"gpu{gid}_{suffix}"
            if cname not in df.columns:
                df[cname] = None
            columns.append(cname)

    df = df[columns]

    # Формат времени: DD-MM-YYYY HH:mm:ss
    try:
        dt = pd.to_datetime(df["timestamp"])
        df["timestamp"] = dt.dt.strftime("%d-%m-%Y %H:%M:%S")
    except Exception:
        pass

    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=args.sheet)
        ws = writer.sheets[args.sheet]

        # Автоширина колонок и перенос строк для автовысоты
        wrap = Alignment(wrap_text=True)
        for col_idx, col_name in enumerate(df.columns, start=1):
            values = [str(col_name)] + ["" if v is None else str(v) for v in df[col_name].tolist()]
            max_len = max(len(v) for v in values)
            # небольшой запас + ограничитель ширины
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = wrap  # Excel сам подстроит высоту при переносе

    print(f"Saved {len(df)} rows to {args.output}")

if __name__ == "__main__":
    main()